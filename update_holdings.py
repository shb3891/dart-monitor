"""
update_holdings.py - 보유내역 자동 업데이트 모듈
- holdings/ 폴더의 최신 xlsx 파일 자동 인식
- ISIN 매칭 → 포트폴리오 시트 보유 정보 갱신
- 신규/사라진 종목 감지 → 텔레봇 알림
- 보유내역_변동이력 시트에 변동 기록
"""

import os
import re
import json
import glob
import time
import requests
import gspread
import pandas as pd
from datetime import datetime, timedelta
from google.oauth2.service_account import Credentials


# ============================================================
# [설정]
# ============================================================
SHEET_ID = os.environ.get('SHEET_ID', '1s73BDNtCPe5mOs9EjBE5npEfcaNtYRyWJxRBUmJI-WA')
TELEGRAM_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN', '')
TELEGRAM_CHAT = os.environ.get('TELEGRAM_CHAT_ID', '')

if not os.environ.get('GCP_SERVICE_ACCOUNT_KEY'):
    raise RuntimeError("GCP_SERVICE_ACCOUNT_KEY 환경변수가 설정되지 않았습니다")

NOW = datetime.utcnow() + timedelta(hours=9)  # KST
TODAY = NOW.strftime('%Y-%m-%d')

HOLDINGS_DIR = 'holdings'  # xlsx 파일 위치

# 시트 이름
SHEET_PORTFOLIO = 0
SHEET_HOLDING_HISTORY = '보유내역_변동이력'


# ============================================================
# [텔레그램 전송]
# ============================================================
def send_telegram(message: str):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT:
        print(f"⚠ 텔레그램 설정 없음")
        return False
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={
                'chat_id': TELEGRAM_CHAT,
                'text': message,
                'parse_mode': 'HTML',
                'disable_web_page_preview': True,
            }, timeout=10
        )
        return r.status_code == 200
    except Exception as e:
        print(f"⚠ 텔레그램 전송 오류: {e}")
        return False


# ============================================================
# [최신 보유내역 파일 찾기]
# ============================================================
def find_latest_holdings_file():
    """holdings/ 폴더에서 가장 최근 수정된 xlsx 파일 반환"""
    if not os.path.isdir(HOLDINGS_DIR):
        print(f"⚠ {HOLDINGS_DIR}/ 폴더가 없음. 폴더 생성 후 xlsx 파일 업로드 필요")
        return None
    
    patterns = [
        os.path.join(HOLDINGS_DIR, '*.xlsx'),
        os.path.join(HOLDINGS_DIR, '*.xls'),
    ]
    files = []
    for p in patterns:
        files.extend(glob.glob(p))
    
    if not files:
        print(f"⚠ {HOLDINGS_DIR}/ 폴더에 xlsx 파일 없음")
        return None
    
    # 수정 시각 기준 최신
    latest = max(files, key=os.path.getmtime)
    mtime = datetime.fromtimestamp(os.path.getmtime(latest))
    print(f"📁 최신 파일: {latest}")
    print(f"   수정일: {mtime.strftime('%Y-%m-%d %H:%M:%S')}")
    return latest


# ============================================================
# [보유내역 xlsx 파싱]
# ============================================================
def parse_holdings_xlsx(filepath):
    """
    xlsx 파일에서 보유 정보 추출 후 ISIN별 집계
    
    Returns:
        dict: {isin: {name, total_amount, total_acq, total_eval, ...}}
    """
    print(f"\n📊 파일 파싱 중: {filepath}")
    
    try:
        df = pd.read_excel(filepath)
    except Exception as e:
        print(f"⚠ 파일 읽기 실패: {e}")
        return {}
    
    # 필요한 컬럼 확인
    required_cols = ['종목코드', '종목명', '수량']
    optional_cols = ['취득가', '시가평가액', '수익률', '발행일자', '상환일자',
                     '예탁원종목코드', '발행기관명', '신용등급']
    
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        print(f"⚠ 필수 컬럼 누락: {missing}")
        return {}
    
    # 종목코드(ISIN) 필터링 (KR로 시작)
    df = df[df['종목코드'].notna()]
    df = df[df['종목코드'].astype(str).str.startswith('KR')]
    
    print(f"  📋 총 {len(df)}행 (ISIN 유효)")
    
    # ISIN별 집계 (한 종목 여러 행이면 합산)
    holdings = {}
    for isin, group in df.groupby('종목코드'):
        # 첫 행 기준 메타 정보
        first = group.iloc[0]
        
        # 합산 (수량, 취득가, 시가평가액)
        total_amount = group['수량'].sum() if '수량' in df.columns else 0
        total_acq = group['취득가'].sum() if '취득가' in df.columns else 0
        total_eval = group['시가평가액'].sum() if '시가평가액' in df.columns else 0
        
        # 수익률은 평균 (또는 가중평균이 더 정확)
        avg_return = group['수익률'].mean() if '수익률' in df.columns else 0
        
        holdings[isin] = {
            'isin': isin,
            'name': str(first.get('종목명', '')).strip(),
            'amount': int(total_amount) if pd.notna(total_amount) else 0,  # 원 단위
            'amount_billion': round(total_amount / 100_000_000, 2),  # 억원
            'acq_price': int(total_acq) if pd.notna(total_acq) else 0,
            'acq_billion': round(total_acq / 100_000_000, 2),
            'eval_price': int(total_eval) if pd.notna(total_eval) else 0,
            'eval_billion': round(total_eval / 100_000_000, 2),
            'return_pct': round(avg_return, 2) if pd.notna(avg_return) else 0,
            'issuer': str(first.get('발행기관명', '')).strip(),
            'rating': str(first.get('신용등급', '')).strip() if pd.notna(first.get('신용등급')) else '',
            'issue_date': str(first.get('발행일자', '')).strip()[:10] if pd.notna(first.get('발행일자')) else '',
            'maturity_date': str(first.get('상환일자', '')).strip()[:10] if pd.notna(first.get('상환일자')) else '',
            'row_count': len(group),  # 몇 개 행에서 왔는지
        }
    
    print(f"  ✅ {len(holdings)}개 고유 종목 (ISIN) 집계 완료")
    print(f"  💰 총 보유금액: {sum(h['amount_billion'] for h in holdings.values()):.0f}억원")
    
    return holdings


# ============================================================
# [포트폴리오 시트 로드 + 컬럼 인덱스 찾기]
# ============================================================
def load_portfolio(ws_portfolio):
    """
    포트폴리오 시트 로드 + 보유관련 컬럼 위치 자동 탐색
    
    Returns:
        (rows, header_map): 시트 데이터 + 컬럼 매핑
    """
    all_values = ws_portfolio.get_all_values()
    if not all_values:
        return [], {}
    
    headers = all_values[0]
    header_map = {}
    
    # 컬럼 자동 인식 (헤더명 패턴 매칭)
    for i, h in enumerate(headers):
        h_clean = h.strip()
        if h_clean in ('종목명', 'A'):
            header_map['name_col'] = i
        elif h_clean in ('ISIN', '채권ISIN', '종목코드'):
            header_map['isin_col'] = i
        elif '보유금액' in h_clean or '보유수량' in h_clean:
            header_map['amount_col'] = i
        elif '취득가' in h_clean and '원' not in h_clean:
            header_map['acq_col'] = i
        elif '평가액' in h_clean or '시가평가' in h_clean:
            header_map['eval_col'] = i
        elif '수익률' in h_clean:
            header_map['return_col'] = i
    
    # ISIN 컬럼이 없으면 B열로 가정 (기존 main.py 구조)
    if 'isin_col' not in header_map:
        header_map['isin_col'] = 1  # B열
    if 'name_col' not in header_map:
        header_map['name_col'] = 0  # A열
    
    rows = all_values[1:]
    print(f"📋 포트폴리오 시트 로드: {len(rows)}개 종목")
    print(f"   컬럼 매핑: {header_map}")
    
    return rows, header_map, headers


# ============================================================
# [신규 보유 컬럼 확보 (없으면 추가)]
# ============================================================
def ensure_holding_columns(ws_portfolio, headers, header_map):
    """
    포트폴리오 시트 끝에 보유 관련 컬럼이 없으면 추가
    추가 컬럼: 보유금액(억원), 취득가(억원), 시가평가액(억원), 수익률(%), 최근업데이트
    
    Returns:
        업데이트된 header_map
    """
    new_cols = []
    
    if 'amount_col' not in header_map:
        new_cols.append('보유금액(억원)')
    if 'acq_col' not in header_map:
        new_cols.append('취득가(억원)')
    if 'eval_col' not in header_map:
        new_cols.append('시가평가액(억원)')
    if 'return_col' not in header_map:
        new_cols.append('수익률(%)')
    
    # 최근업데이트 컬럼도 추가
    if '최근업데이트' not in headers:
        new_cols.append('최근업데이트')
    
    if not new_cols:
        print("  ✅ 보유 관련 컬럼 이미 존재")
        return header_map
    
    # 마지막 컬럼 인덱스
    start_col = len(headers)
    
    # 헤더 추가
    print(f"  🆕 신규 컬럼 추가: {new_cols}")
    
    from openpyxl.utils import get_column_letter
    start_letter = get_column_letter(start_col + 1)
    end_letter = get_column_letter(start_col + len(new_cols))
    
    range_str = f'{start_letter}1:{end_letter}1'
    
    # 시트 컬럼 수 확인 후 부족하면 확장
    current_cols = ws_portfolio.col_count
    needed_cols = start_col + len(new_cols)
    if needed_cols > current_cols:
        add_cols = needed_cols - current_cols + 2  # 여유있게 2개 더
        print(f"  🔧 시트 컬럼 확장: {current_cols} → {current_cols + add_cols}")
        ws_portfolio.add_cols(add_cols)
        time.sleep(1.0)
    
    ws_portfolio.update([new_cols], range_name=range_str)
    
    # 스타일링
    ws_portfolio.format(range_str, {
        'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
        'backgroundColor': {'red': 0.4, 'green': 0.5, 'blue': 0.3},
        'horizontalAlignment': 'CENTER',
    })
    time.sleep(1.0)
    
    # header_map 업데이트
    new_col_idx = start_col
    for col_name in new_cols:
        if '보유금액' in col_name:
            header_map['amount_col'] = new_col_idx
        elif '취득가' in col_name:
            header_map['acq_col'] = new_col_idx
        elif '평가액' in col_name:
            header_map['eval_col'] = new_col_idx
        elif '수익률' in col_name:
            header_map['return_col'] = new_col_idx
        elif '최근업데이트' in col_name:
            header_map['updated_col'] = new_col_idx
        new_col_idx += 1
    
    return header_map


# ============================================================
# [기존 vs 새 보유 비교]
# ============================================================
def compare_holdings(portfolio_rows, header_map, new_holdings):
    """
    포트폴리오 시트의 기존 보유 vs 새 보유내역 비교
    
    Returns:
        {
            'matched': [(row_idx, isin, old_amount, new_data), ...],
            'new_isins': [isin, ...],     # 보유내역에만 있음 (신규)
            'gone_isins': [isin, ...],     # 포트폴리오에만 있음 (매도/상환?)
        }
    """
    isin_col = header_map.get('isin_col', 1)
    amount_col = header_map.get('amount_col')
    
    # 포트폴리오의 ISIN 집합
    portfolio_isins = {}
    for i, row in enumerate(portfolio_rows):
        if len(row) > isin_col and row[isin_col].strip().startswith('KR'):
            isin = row[isin_col].strip()
            old_amount = 0
            if amount_col is not None and len(row) > amount_col:
                try:
                    old_amount = float(row[amount_col].replace(',', '')) if row[amount_col].strip() else 0
                except:
                    old_amount = 0
            portfolio_isins[isin] = {
                'row_idx': i + 2,  # 시트의 실제 행 번호 (1-indexed + 헤더)
                'name': row[header_map.get('name_col', 0)] if len(row) > header_map.get('name_col', 0) else '',
                'old_amount': old_amount,
            }
    
    matched = []
    new_isins = []
    gone_isins = []
    
    # 신규/매칭
    for isin, new_data in new_holdings.items():
        if isin in portfolio_isins:
            matched.append({
                'row_idx': portfolio_isins[isin]['row_idx'],
                'isin': isin,
                'name': portfolio_isins[isin]['name'],
                'old_amount': portfolio_isins[isin]['old_amount'],
                'new': new_data,
            })
        else:
            new_isins.append(isin)
    
    # 사라진 종목 (포트폴리오에는 있는데 보유내역에 없음)
    # 단, 기존에 amount > 0 였던 경우만 (이미 0이면 알 필요 없음)
    for isin, p_data in portfolio_isins.items():
        if isin not in new_holdings and p_data['old_amount'] > 0:
            gone_isins.append({
                'isin': isin,
                'name': p_data['name'],
                'row_idx': p_data['row_idx'],
                'old_amount': p_data['old_amount'],
            })
    
    print(f"\n📊 비교 결과:")
    print(f"  ✅ 매칭: {len(matched)}개")
    print(f"  🆕 신규: {len(new_isins)}개")
    print(f"  ⚠️ 사라짐: {len(gone_isins)}개")
    
    return {
        'matched': matched,
        'new_isins': new_isins,
        'gone_isins': gone_isins,
    }


# ============================================================
# [포트폴리오 시트 업데이트]
# ============================================================
def update_portfolio_sheet(ws_portfolio, header_map, matched, new_holdings):
    """
    매칭된 종목의 보유 정보 업데이트
    """
    print(f"\n📝 포트폴리오 시트 업데이트 중 ({len(matched)}개 종목)...")
    
    amount_col = header_map.get('amount_col')
    acq_col = header_map.get('acq_col')
    eval_col = header_map.get('eval_col')
    return_col = header_map.get('return_col')
    updated_col = header_map.get('updated_col')
    
    from openpyxl.utils import get_column_letter
    
    # 컬럼별로 일괄 업데이트 (효율적)
    updates_by_col = {}
    
    for m in matched:
        row_idx = m['row_idx']
        new = m['new']
        
        if amount_col is not None:
            updates_by_col.setdefault(amount_col, []).append((row_idx, new['amount_billion']))
        if acq_col is not None:
            updates_by_col.setdefault(acq_col, []).append((row_idx, new['acq_billion']))
        if eval_col is not None:
            updates_by_col.setdefault(eval_col, []).append((row_idx, new['eval_billion']))
        if return_col is not None:
            updates_by_col.setdefault(return_col, []).append((row_idx, new['return_pct']))
        if updated_col is not None:
            updates_by_col.setdefault(updated_col, []).append((row_idx, TODAY))
    
    # 컬럼별로 batch update
    for col_idx, cells in updates_by_col.items():
        col_letter = get_column_letter(col_idx + 1)
        # 정렬 후 연속된 범위로 묶기
        cells.sort(key=lambda x: x[0])
        
        # 한 번에 update (range를 사용)
        if not cells:
            continue
        
        # 행 번호 범위
        min_row = cells[0][0]
        max_row = cells[-1][0]
        
        # 값 배열 (빈 행은 빈 문자열)
        cell_dict = {r: v for r, v in cells}
        values = []
        for r in range(min_row, max_row + 1):
            values.append([cell_dict.get(r, '')])
        
        range_str = f'{col_letter}{min_row}:{col_letter}{max_row}'
        try:
            ws_portfolio.update(values, range_name=range_str)
            time.sleep(0.5)
        except Exception as e:
            print(f"  ⚠ 컬럼 {col_letter} 업데이트 실패: {e}")
    
    print(f"  ✅ {len(matched)}개 종목 업데이트 완료")


def add_new_isins_to_portfolio(ws_portfolio, header_map, new_isins, new_holdings):
    """
    신규 ISIN을 포트폴리오 시트에 추가
    """
    if not new_isins:
        return
    
    print(f"\n🆕 신규 종목 {len(new_isins)}개 추가 중...")
    
    # 현재 데이터 끝 행
    all_values = ws_portfolio.get_all_values()
    next_row = len(all_values) + 1
    
    # 헤더 컬럼 수
    headers = all_values[0] if all_values else []
    total_cols = len(headers)
    
    isin_col = header_map.get('isin_col', 1)
    name_col = header_map.get('name_col', 0)
    amount_col = header_map.get('amount_col')
    acq_col = header_map.get('acq_col')
    eval_col = header_map.get('eval_col')
    return_col = header_map.get('return_col')
    updated_col = header_map.get('updated_col')
    
    rows_to_add = []
    for isin in new_isins:
        new = new_holdings[isin]
        row = [''] * total_cols
        
        if name_col < total_cols:
            row[name_col] = new['name']
        if isin_col < total_cols:
            row[isin_col] = isin
        if amount_col is not None and amount_col < total_cols:
            row[amount_col] = new['amount_billion']
        if acq_col is not None and acq_col < total_cols:
            row[acq_col] = new['acq_billion']
        if eval_col is not None and eval_col < total_cols:
            row[eval_col] = new['eval_billion']
        if return_col is not None and return_col < total_cols:
            row[return_col] = new['return_pct']
        if updated_col is not None and updated_col < total_cols:
            row[updated_col] = TODAY
        
        rows_to_add.append(row)
    
    if rows_to_add:
        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(total_cols)
        range_str = f'A{next_row}:{last_col_letter}{next_row + len(rows_to_add) - 1}'
        
        try:
            ws_portfolio.update(rows_to_add, range_name=range_str)
            time.sleep(1.0)
            
            # 노란색 강조 (신규 표시)
            ws_portfolio.format(range_str, {
                'backgroundColor': {'red': 1, 'green': 0.95, 'blue': 0.8}
            })
            time.sleep(0.5)
            print(f"  ✅ {len(rows_to_add)}개 추가 완료 (노란색 표시)")
        except Exception as e:
            print(f"  ⚠ 추가 실패: {e}")


def mark_gone_isins(ws_portfolio, header_map, gone_isins):
    """
    사라진 종목의 보유금액을 0으로 (매도/상환 표시)
    """
    if not gone_isins:
        return
    
    print(f"\n⚠️ 사라진 종목 {len(gone_isins)}개 처리 중...")
    
    amount_col = header_map.get('amount_col')
    eval_col = header_map.get('eval_col')
    updated_col = header_map.get('updated_col')
    
    if amount_col is None:
        return
    
    from openpyxl.utils import get_column_letter
    
    for gone in gone_isins:
        row_idx = gone['row_idx']
        
        # 보유금액 0으로
        if amount_col is not None:
            col_letter = get_column_letter(amount_col + 1)
            try:
                ws_portfolio.update([[0]], range_name=f'{col_letter}{row_idx}')
                time.sleep(0.3)
            except Exception as e:
                print(f"  ⚠ {gone['name']} 업데이트 실패: {e}")
        
        # 시가평가액 0
        if eval_col is not None:
            col_letter = get_column_letter(eval_col + 1)
            try:
                ws_portfolio.update([[0]], range_name=f'{col_letter}{row_idx}')
                time.sleep(0.3)
            except Exception as e:
                pass
        
        # 회색 표시
        try:
            ws_portfolio.format(f'A{row_idx}:Z{row_idx}', {
                'backgroundColor': {'red': 0.92, 'green': 0.92, 'blue': 0.92}
            })
            time.sleep(0.3)
        except Exception as e:
            pass
    
    print(f"  ✅ {len(gone_isins)}개 처리 완료 (회색 표시)")


# ============================================================
# [변동 이력 기록]
# ============================================================
def record_change_history(sh, matched, new_isins, gone_isins, new_holdings):
    """
    보유내역_변동이력 시트에 변동 사항 기록
    """
    try:
        ws = sh.worksheet(SHEET_HOLDING_HISTORY)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=SHEET_HOLDING_HISTORY, rows=500, cols=7)
        headers = ['기록일자', '종목명', 'ISIN', '변동유형', '이전 금액(억)', '현재 금액(억)', '차이(억)']
        ws.update([headers], range_name='A1:G1')
        ws.format('A1:G1', {
            'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
            'backgroundColor': {'red': 0.4, 'green': 0.3, 'blue': 0.5},
            'horizontalAlignment': 'CENTER',
        })
        time.sleep(1.0)
        print(f"  ✅ '{SHEET_HOLDING_HISTORY}' 시트 새로 생성")
    
    # 새 행들 작성
    rows = []
    
    # 신규
    for isin in new_isins:
        new = new_holdings[isin]
        rows.append([TODAY, new['name'], isin, '🆕 신규', 0, new['amount_billion'], 
                     new['amount_billion']])
    
    # 사라진 종목
    for gone in gone_isins:
        rows.append([TODAY, gone['name'], gone['isin'], '⚠️ 매도/상환', 
                     gone['old_amount'], 0, -gone['old_amount']])
    
    # 매칭 - 큰 변동 (10% 이상)만 기록
    for m in matched:
        old = m['old_amount']
        new_amt = m['new']['amount_billion']
        if old == 0 and new_amt == 0:
            continue
        # 변동률 계산
        if old > 0:
            change_pct = abs((new_amt - old) / old) * 100
            if change_pct < 10:  # 10% 미만은 기록 안 함
                continue
        diff = round(new_amt - old, 2)
        if abs(diff) < 1:  # 1억 미만 변동은 기록 안 함
            continue
        change_type = '📈 증가' if diff > 0 else '📉 감소'
        rows.append([TODAY, m['name'], m['isin'], change_type, old, new_amt, diff])
    
    if not rows:
        print("  ℹ 기록할 변동 없음")
        return
    
    # 기존 데이터 + 신규 (역순으로 위에 쌓이게)
    existing = ws.get_all_values()
    existing_data = existing[1:] if len(existing) > 1 else []
    
    all_rows = rows + existing_data  # 신규가 위로
    
    # 업데이트
    ws.batch_clear([f'A2:G500'])
    time.sleep(0.5)
    if all_rows:
        ws.update(all_rows, range_name=f'A2:G{len(all_rows)+1}')
        time.sleep(1.0)
    print(f"  ✅ 변동이력 {len(rows)}건 신규 기록 (총 {len(all_rows)}건)")


# ============================================================
# [텔레그램 알림]
# ============================================================
def send_alerts(matched, new_isins, gone_isins, new_holdings):
    """변동 사항을 텔레봇 알림으로 전송"""
    
    # 1. 신규 종목 알림 (개별)
    for isin in new_isins:
        new = new_holdings[isin]
        msg = (
            f"🆕 <b>신규 보유 종목 추가</b>\n\n"
            f"🔹 <b>{new['name']}</b>\n"
            f"🆔 <code>{isin}</code>\n"
            f"💰 보유: <b>{new['amount_billion']}억원</b>\n"
            f"💵 취득가: {new['acq_billion']}억원\n"
            f"📊 평가액: {new['eval_billion']}억원\n"
            f"📈 수익률: {new['return_pct']}%\n"
            + (f"🏢 발행사: {new['issuer']}\n" if new['issuer'] else '')
            + (f"📅 발행: {new['issue_date']} ~ {new['maturity_date']}\n" if new['issue_date'] else '')
            + f"\n⚠️ 포트폴리오 시트에 자동 추가됨"
        )
        send_telegram(msg)
        time.sleep(0.5)
    
    # 2. 사라진 종목 알림 (개별)
    for gone in gone_isins:
        msg = (
            f"⚠️ <b>보유 종목 사라짐</b>\n\n"
            f"🔹 <b>{gone['name']}</b>\n"
            f"🆔 <code>{gone['isin']}</code>\n"
            f"💰 이전 보유: {gone['old_amount']}억원\n"
            f"💰 현재 보유: 0억원\n\n"
            f"❓ 매도 또는 상환 여부 확인 필요"
        )
        send_telegram(msg)
        time.sleep(0.5)
    
    # 3. 처리 완료 요약
    total = len(matched) + len(new_isins)
    total_amount = sum(new_holdings[isin]['amount_billion'] for isin in new_holdings)
    
    # 큰 변동 카운트
    big_changes = []
    for m in matched:
        old = m['old_amount']
        new_amt = m['new']['amount_billion']
        if old > 0:
            change_pct = abs((new_amt - old) / old) * 100
            if change_pct >= 10:
                big_changes.append(m)
    
    summary = (
        f"📊 <b>보유내역 업데이트 완료</b>\n\n"
        f"📅 기준일: {TODAY}\n"
        f"📌 총 보유: <b>{total}개 종목 / {total_amount:.0f}억원</b>\n\n"
        f"✅ 매칭: {len(matched)}개\n"
        f"🆕 신규: {len(new_isins)}개\n"
        f"⚠️ 매도/상환: {len(gone_isins)}개\n"
        f"📈📉 큰 변동(±10%): {len(big_changes)}개"
    )
    send_telegram(summary)
    print(f"\n📨 알림 전송 완료")


# ============================================================
# [메인]
# ============================================================
def main():
    print(f"🤖 보유내역 자동 업데이트 시작 ({TODAY})")
    
    # 1. 최신 파일 찾기
    filepath = find_latest_holdings_file()
    if not filepath:
        print("⚠ 처리할 파일 없음. 종료.")
        return
    
    # 2. 파일 파싱
    new_holdings = parse_holdings_xlsx(filepath)
    if not new_holdings:
        print("⚠ 파싱 실패. 종료.")
        return
    
    # 3. Google Sheets 연결
    creds_json = json.loads(os.environ.get('GCP_SERVICE_ACCOUNT_KEY'))
    creds = Credentials.from_service_account_info(creds_json, scopes=[
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive',
    ])
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SHEET_ID)
    ws_portfolio = sh.get_worksheet(SHEET_PORTFOLIO)
    
    # 4. 포트폴리오 로드 + 컬럼 확보
    portfolio_rows, header_map, headers = load_portfolio(ws_portfolio)
    header_map = ensure_holding_columns(ws_portfolio, headers, header_map)
    
    # 5. 비교
    comparison = compare_holdings(portfolio_rows, header_map, new_holdings)
    
    # 6. 업데이트
    update_portfolio_sheet(ws_portfolio, header_map, 
                          comparison['matched'], new_holdings)
    add_new_isins_to_portfolio(ws_portfolio, header_map, 
                                comparison['new_isins'], new_holdings)
    mark_gone_isins(ws_portfolio, header_map, comparison['gone_isins'])
    
    # 7. 변동 이력 기록
    record_change_history(sh, comparison['matched'], 
                          comparison['new_isins'], 
                          comparison['gone_isins'], 
                          new_holdings)
    
    # 8. 텔레봇 알림
    send_alerts(comparison['matched'], comparison['new_isins'],
                comparison['gone_isins'], new_holdings)
    
    print(f"\n🏁 완료!")


if __name__ == '__main__':
    main()
