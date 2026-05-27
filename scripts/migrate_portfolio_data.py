"""
migrate_portfolio_data.py — 엑셀 권리현황 → 포트폴리오 시트 데이터 마이그레이션

작업:
1. 권리현황 엑셀(20260430)에서 각 박스 파싱
2. 113개 펀드 종목과 매칭 (별칭 사전 활용)
3. 매칭된 종목의 정보를 포트폴리오 시트에 자동 입력:
   - E COUPON
   - F YTM
   - G 행사가액 (숫자만)
   - H 리픽싱플로어 (% 추출)
   - I CALL비율
   - J YTC
   - Q YTP
   - T PUT주기
   - U 다음리픽싱예정일
   - V 전환청구시작일
   - W 전환청구종료일
4. 매칭 안 된 종목은 빈칸으로 두고 텔레로 알림 가능

사용법:
- 엑셀 파일을 dart-monitor 레포 루트에 업로드 (예: '권리현황_20260430.xlsx')
- GitHub Actions 수동 실행
"""

import os
import json
import re
from datetime import datetime
import openpyxl
import gspread
from google.oauth2.service_account import Credentials


SHEET_ID = os.environ.get('SHEET_ID', '1s73BDNtCPe5mOs9EjBE5npEfcaNtYRyWJxRBUmJI-WA')
EXCEL_FILE = os.environ.get('EXCEL_FILE', '권리현황_20260430.xlsx')
EXCEL_SHEET = os.environ.get('EXCEL_SHEET', '권리현황 20260430')


# ============================================================
# [엑셀 표기 → 펀드 표기 별칭]
# ============================================================
ALIASES = {
    'KBI메탈': '케이비아이메탈',
    '오로라': '오로라월드',
    'KH바텍': '케이에이치바텍',
    'DYP': '디와이피',
    'KG에코솔루션': '케이지에코솔루션',
    '엘앤케이바이오': '엘앤케이바이오메드',
    'NEW': '넥스트엔터테인먼트월드',
    'SCL사이언스': '에스씨엘사이언스',
    'TS인베스트먼트': '티에스인베스트먼트',
    '에르코스': '에르코스농업회사법인',
}


# ============================================================
# [Google Sheets 연결]
# ============================================================
creds_json = json.loads(os.environ.get('GCP_SERVICE_ACCOUNT_KEY'))
creds = Credentials.from_service_account_info(creds_json, scopes=[
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive',
])
gc = gspread.authorize(creds)
sh = gc.open_by_key(SHEET_ID)


# ============================================================
# [유틸: 종목명 추출 / 정규화]
# ============================================================
def extract_bond_name(raw):
    """엑셀 셀에서 종목명만 깔끔하게 추출."""
    if not raw:
        return ''
    s = str(raw).strip()
    s = s.split('\n')[0].strip()
    s = re.sub(r'\([^)]*\)', '', s).strip()
    s = re.split(r'\s*[-–]\s*', s)[0].strip()
    s = re.split(r'\s*/\s*\d', s)[0].strip()
    m = re.match(r'^(.+?(?:CB|EB|BW))\s*', s)
    if m:
        return m.group(1).strip()
    return s


def normalize(s):
    return re.sub(r'\s+', '', s).strip()


def apply_alias(name):
    for ex, fund in ALIASES.items():
        if name.startswith(ex):
            return name.replace(ex, fund, 1)
    return name


# ============================================================
# [유틸: 데이터 파싱]
# ============================================================
def parse_rate(val):
    """이율 값 파싱.
    
    예: 0 → '0%'
        0.02 → '2%'
        '2% + 가산금리' → '2% + 가산금리'
    """
    if val is None or val == '':
        return ''
    if isinstance(val, (int, float)):
        if val == 0:
            return '0%'
        # 0.02 같은 소수면 백분율 변환
        if val < 1:
            return f'{val*100:.2f}%'.rstrip('0').rstrip('.') + '%' if val*100 != int(val*100) else f'{int(val*100)}%'
        return f'{val}%'
    return str(val).strip()


def parse_strike_price(val):
    """행사가액 파싱: '4,470원(2026.01.22 변경) → 2026.09.22 변경예정' → 4470"""
    if not val:
        return '', ''
    s = str(val).strip()
    
    # 가격 추출 (첫 번째 숫자,숫자원 패턴)
    price_m = re.search(r'([\d,]+)\s*원', s)
    price = ''
    if price_m:
        price = price_m.group(1).replace(',', '')
    
    # 변경예정일 추출
    next_change = ''
    next_m = re.search(r'(\d{4})\.(\d{1,2})\.(\d{1,2})\s*변경예정', s)
    if next_m:
        next_change = f"{next_m.group(1)}-{next_m.group(2).zfill(2)}-{next_m.group(3).zfill(2)}"
    
    return price, next_change


def parse_refixing(val):
    """리픽싱 한도 파싱.
    
    예: '70% (6,150원 → 4,305원), 8개월 단위' → '70%'
        '실적 리픽싱 70%' → '70%'
        'IPO 공모가 70%' → '70%'
    """
    if not val:
        return ''
    s = str(val).strip()
    m = re.search(r'(\d+(?:\.\d+)?)\s*%', s)
    if m:
        return f"{m.group(1)}%"
    return s


def parse_put_cycle_ytp(val):
    """풋 주기 + YTP 파싱.
    
    예: '3개월 단위, YTP 2%' → ('3개월', '2%')
        '3개월 단위, YTP 0%' → ('3개월', '0%')
    """
    if not val:
        return '', ''
    s = str(val).strip()
    
    cycle = ''
    cycle_m = re.search(r'(\d+개월)\s*단위', s)
    if cycle_m:
        cycle = cycle_m.group(1)
    
    ytp = ''
    ytp_m = re.search(r'YTP\s*([\d.]+)\s*%', s)
    if ytp_m:
        ytp = f"{ytp_m.group(1)}%"
    
    return cycle, ytp


def parse_call_ratio(val):
    """콜 비율 파싱.
    
    예: '콜옵션 없음' → '없음'
        '콜 행사완료' → '행사완료'
        '20%' → '20%'
        숫자 → 그대로
    """
    if not val:
        return ''
    s = str(val).strip()
    if '없음' in s:
        return '없음'
    if '행사완료' in s or '완료' in s:
        return '행사완료'
    m = re.search(r'(\d+(?:\.\d+)?)\s*%', s)
    if m:
        return f"{m.group(1)}%"
    return s


def parse_date(val):
    """날짜 파싱.
    
    예: '2026.02.27' → '2026-02-27'
        '2026.02.27 ~ 2026.03.30' → '' (범위는 풋콜스케줄에서 처리)
    """
    if not val:
        return ''
    s = str(val).strip()
    # 범위가 아닌 단일 날짜만
    if '~' in s:
        return ''
    m = re.match(r'(\d{4})\.(\d{1,2})\.(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    return ''


# ============================================================
# [엑셀 박스 파싱]
# ============================================================
def extract_box_data(ws, header_row):
    """한 박스에서 col 4, 6, 8, 10 각각의 종목 정보 추출."""
    bonds = []
    for col in [4, 6, 8, 10]:
        bond_raw = ws.cell(row=header_row, column=col).value
        if not bond_raw:
            continue
        
        name = extract_bond_name(bond_raw)
        if not name:
            continue
        
        rel = lambda offset: ws.cell(row=header_row + offset, column=col).value
        
        info = {
            'name': name,
            'right_begin': rel(6),    # 권리청구 시작
            'right_end': rel(7),      # 권리청구 종료
            'coupon': rel(4),         # 표면 이율
            'ytm': rel(5),            # 만기 이율
            'strike': rel(8),         # 행사가액
            'refixing': rel(9),       # Refixing
            'put_cycle_ytp': rel(11), # 풋 주기/YTP
            'call_ratio': rel(13),    # 콜 비율
            # YTC는 콜 비율 영역에 따로 없음 (대부분 빈칸이나 콜이 있는 경우 별도 추출)
        }
        bonds.append(info)
    return bonds


def parse_excel(excel_path, sheet_name):
    """엑셀 파일에서 모든 박스 파싱 → 종목명 → 데이터 dict 반환."""
    print(f"📂 엑셀 로드: {excel_path} [{sheet_name}]")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    
    # 박스 헤더 행 모두 찾기
    box_headers = []
    for r in range(1, ws.max_row + 1):
        c2 = ws.cell(row=r, column=2).value
        if c2 and '종목명' == str(c2).strip():
            box_headers.append(r)
    print(f"  ✅ 박스 {len(box_headers)}개 발견")
    
    # 종목별 데이터 추출
    bonds = {}  # normalized_name → bond_info
    for header_row in box_headers:
        box_bonds = extract_box_data(ws, header_row)
        for b in box_bonds:
            key = normalize(b['name'])
            bonds[key] = b
    
    print(f"  ✅ 총 {len(bonds)}개 종목 데이터 추출")
    return bonds


# ============================================================
# [메인]
# ============================================================
def main():
    today = datetime.now().strftime('%Y-%m-%d')
    print(f"🚀 포트폴리오 데이터 마이그레이션 시작 ({today})\n")
    
    # 1. 엑셀 파싱
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ 엑셀 파일 없음: {EXCEL_FILE}")
        print(f"   dart-monitor 레포 루트에 파일 업로드 필요")
        return
    
    excel_bonds = parse_excel(EXCEL_FILE, EXCEL_SHEET)
    
    # 별칭 적용해서 추가 매핑
    for k in list(excel_bonds.keys()):
        info = excel_bonds[k]
        aliased = apply_alias(info['name'])
        if aliased != info['name']:
            excel_bonds[normalize(aliased)] = info
    
    # 2. 포트폴리오 시트 로드
    print(f"\n📋 포트폴리오 시트 로드...")
    try:
        ws_port = sh.worksheet('포트폴리오')
    except Exception as e:
        print(f"❌ 포트폴리오 시트 없음: {e}")
        print(f"   먼저 setup_portfolio_sheet.py 실행 필요")
        return
    
    port_rows = ws_port.get_all_values()
    print(f"  ✅ 포트폴리오 종목: {len(port_rows) - 1}개")
    
    # 3. 종목별 매칭 + 데이터 추출
    print(f"\n🔍 113개 종목 매칭 + 데이터 추출 중...")
    updates = []   # (row_idx, [E, F, G, H, I, J, Q, T, U, V, W])
    matched_count = 0
    unmatched = []
    
    for i, row in enumerate(port_rows[1:], start=2):
        if len(row) < 2 or not row[0].startswith('KR'):
            continue
        
        isin = row[0].strip()
        bond_name = row[1].strip()
        normalized = normalize(bond_name)
        
        # 엑셀에서 찾기
        excel_info = excel_bonds.get(normalized)
        if not excel_info:
            unmatched.append((isin, bond_name))
            continue
        
        # 데이터 추출
        coupon = parse_rate(excel_info['coupon'])
        ytm = parse_rate(excel_info['ytm'])
        strike, next_refix = parse_strike_price(excel_info['strike'])
        refixing = parse_refixing(excel_info['refixing'])
        call_ratio = parse_call_ratio(excel_info['call_ratio'])
        put_cycle, ytp = parse_put_cycle_ytp(excel_info['put_cycle_ytp'])
        right_begin = parse_date(excel_info['right_begin'])
        right_end = parse_date(excel_info['right_end'])
        
        # 컬럼 매핑: E F G H I J / Q T U / V W
        # (J=YTC, R/S=현재CALL — 일단 빈칸)
        updates.append({
            'row': i,
            'E': coupon,
            'F': ytm,
            'G': strike,
            'H': refixing,
            'I': call_ratio,
            'J': '',          # YTC (엑셀에 명시적으로 없음)
            'Q': ytp,
            'T': put_cycle,
            'U': next_refix,
            'V': right_begin,
            'W': right_end,
        })
        matched_count += 1
    
    print(f"  ✅ 매칭 성공: {matched_count}/{len(port_rows) - 1}개")
    if unmatched:
        print(f"  ⚠ 매칭 실패: {len(unmatched)}개")
        for isin, name in unmatched:
            print(f"     - {isin} {name}")
    
    # 4. 시트 업데이트 (batch)
    print(f"\n📝 시트 업데이트 중...")
    
    # E:J 컬럼 묶음
    ef_updates = []
    qq_updates = []
    tu_updates = []
    vw_updates = []
    
    for u in updates:
        ef_updates.append({
            'range': f"E{u['row']}:J{u['row']}",
            'values': [[u['E'], u['F'], u['G'], u['H'], u['I'], u['J']]],
        })
        qq_updates.append({
            'range': f"Q{u['row']}",
            'values': [[u['Q']]],
        })
        tu_updates.append({
            'range': f"T{u['row']}:U{u['row']}",
            'values': [[u['T'], u['U']]],
        })
        vw_updates.append({
            'range': f"V{u['row']}:W{u['row']}",
            'values': [[u['V'], u['W']]],
        })
    
    # 50개씩 묶어서 batch update
    def batch_update_chunks(updates_list, label):
        BATCH_SIZE = 50
        for i in range(0, len(updates_list), BATCH_SIZE):
            chunk = updates_list[i:i + BATCH_SIZE]
            ws_port.batch_update(chunk)
            print(f"  ✅ {label}: {min(i + BATCH_SIZE, len(updates_list))}/{len(updates_list)}개")
    
    batch_update_chunks(ef_updates, 'E~J 컬럼 (COUPON/YTM/행사가액/리픽싱/CALL비율/YTC)')
    batch_update_chunks(qq_updates, 'Q 컬럼 (YTP)')
    batch_update_chunks(tu_updates, 'T~U 컬럼 (PUT주기/다음리픽싱)')
    batch_update_chunks(vw_updates, 'V~W 컬럼 (전환청구기간)')
    
    print(f"\n🏁 완료!")
    print(f"  📊 매칭 성공: {matched_count}/{len(port_rows) - 1}개")
    print(f"  👉 https://docs.google.com/spreadsheets/d/{SHEET_ID}")


if __name__ == '__main__':
    main()
